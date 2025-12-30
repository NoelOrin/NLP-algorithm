#coding='utf-8'
import openpyxl  # 替换xlrd以支持xlsx文件
import re
import requests
import xlwt
import os
import time as t
import random
import numpy as np	
import datetime
import urllib3
import json
import sys
import argparse
from multiprocessing.dummy import Pool as ThreadPool

urllib3.disable_warnings()

# 全局变量
cookie = ''
headers = {}

def init_crawler(config_file=None):
    """初始化爬虫配置"""
    global cookie, headers
    
    # 如果提供了配置文件，从文件读取cookie
    if config_file and os.path.exists(config_file):
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                cookie = config.get('cookie', '')
                print(f"从配置文件加载cookie，长度: {len(cookie)}")
        except Exception as e:
            print(f"读取配置文件失败: {e}")
            cookie = ''
    
    # 如果没有cookie，使用默认值
    if not cookie:
        cookie = 'SCF=ArMQAAJmyIG49ARh17nK3PgjUV15xWAqieE_jmgpqmK6SjDbtkttYBrQnW5kx98k1R8HC7q0dJ9FfZb3KVH_1ys.; SINAGLOBAL=4456018209732.372.1766759621315; ULV=1766759621322:1:1:1:4456018209732.372.1766759621315:; WBPSESS=Dt2hbAUaXfkVprjyrAZT_MS6MXUtjhrWrCoRAV-x2glk7JcacDVexL9A5bLFu1UZI33sB7OGiMOlSfP8tlAFJsJCWvzeIoZzjkHs105tepKR_NjXlQNrhwyMxB0lSCojpGUPgyZBQEiOYGaK6UUWzvHYIxAeFdsHh1WcOSU4WqkSu55uBpL9_iYUj0wJ-8d1sIyH66GkfKeTXJI00MD7yw==; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9Wh8HILbCB1TnVb5veiMiHfR5JpX5KzhUgL.Fo.7eo-Reh54e0e2dJLoIp7LxKML1KBLBKnLxKqL1hnLBoM4ehzf1h571Ke0; SUB=_2A25ESwjCDeRhGe9O6VcZ8C7FyD-IHXVnKQQKrDV8PUNbmtANLUHgkW9NdKWo_WYz_OvhpyYUPCTsqNWlPeiSm1hg; ALF=02_1769407890; XSRF-TOKEN=gtewfdUNL6qOfA71wqJy7rc0'
        print("使用默认cookie")
    
    # 设置headers
    headers = {
        'Accept-Encoding': 'gzip, deflate, sdch',
        'Accept-Language': 'en-US,en;q=0.8',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Referer': 'https://www.baidu.com/',
        'Connection': 'keep-alive',
        'Cookie': cookie,
    }
    
    print("爬虫初始化完成")

def parse_arguments():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description='微博评论爬虫')
    parser.add_argument('--config', type=str, help='配置文件路径')
    parser.add_argument('--cookie', type=str, help='直接传入cookie')
    return parser.parse_args()

def require(url):
	"""获取网页源码"""
	while True:
		try:
			response = requests.get(url, headers=headers,timeout=(30,50),verify=False)
			#print(url)
			code_1=response.status_code
			#print(type(code_1))
			t.sleep(random.randint(1,2))
			if code_1==200:
				print('正常爬取中，状态码：'+str(code_1))#状态码
				t.sleep(random.randint(1,2))
				break
			else:
				print('请求异常，重试中，状态码为：'+str(code_1))#状态码
				t.sleep(random.randint(2,3))
				continue
		except:
			t.sleep(random.randint(2,3))
			continue

	#print(response.encoding)#首选编码
	#response.encoding=response.apparent_encoding
	html=response.text#源代码文本
	return html

def html_1(url):#返回网页源码和评论页数
	html=require(url)
	try:
		page=re.findall('&nbsp;1/(.*?)页',html,re.S)
		page=int(page[0])
	except:
		page=0
	#page=re.findall('<input name="mp" type="hidden" value="(.*?)">',html,re.S)
	return html,page

def count(alls):
	n=0
	for all in alls:
		for i in all:
			n=n+1
	return n


def body(h_1):#主体
	html_2=re.findall('<div class="c" id="C.*?">(.*?)</div>',str(h_1),re.S)
	html_2=str(html_2)

	user_ids=re.findall('<a href=".*?&amp;fuid=(.*?)&amp;.*?">举报</a> ',html_2,re.S)#从举报链接入手

	names_0=re.findall('<a href=.*?>(.*?)</a>',html_2,re.S)
	names=[]#用户名
	ma=[ '举报', '赞[]', '回复']
	pattern = re.compile(r'\d+')#匹配数字
	for i in names_0:
		i=re.sub(pattern, "", i)
		if i not in ma:
			if '@' not in i:
				names.append(i)

	pattern_0= re.compile(r'回复<a href=.*?</a>:')#匹配回复前缀
	pattern_0_1= re.compile(r'<a href=.*?</a>')#匹配回复内容后面的表情图片地址
	pattern_0_2= re.compile(r'<img alt=.*?/>')#匹配回复内容的图片地址
	contents=[]#评论内容
	contents_2=[]#评论内容初步
	contents_0=re.findall('<span class="ctt">(.*?)</span>',html_2,re.S)#一级
	contents_1=re.findall('<a href=.*?>@.*?</a>(.*?)<a href=.*?>举报</a> ',html_2,re.S)#二级

	for i in contents_0:
		i=re.sub(pattern_0,'',i)
		i=re.sub(pattern_0_1,'',i)
		i=re.sub(pattern_0_2,'',i)
		i=i.replace(':','')
		i=i.strip()
		contents_2.append(i)

	for i in contents_1:
		i=re.sub(pattern_0,'',i)
		i=re.sub(pattern_0_1,'',i)
		i=re.sub(pattern_0_2,'',i)
		i=i.replace('</span>','')
		i=i.replace('&nbsp;','')
		i=i.replace(':','')
		i=i.strip()
		contents_2.append(i)

	for i in contents_2:
		i=re.sub(r'\s','',i)#去除空白
		if len(i)==0:
			pass
		else:
			contents.append(i)
	times_0=re.findall('<span class="ct">(.*?)</span>',html_2,re.S)
	times=[]#时间
	pattern_1= re.compile(r'\d{2}月\d{2}日')#匹配日期
	for i in times_0:
		try:
			t_1= re.match(pattern_1, i).group()
		except:
			a=datetime.datetime.now().strftime('%m%d')
			t_1=a[:2]+'月'+a[2:]+'日'#改为当天
		times.append(t_1)

	all=[]
	for i in range(len(user_ids)):
		try:
			al=[user_ids[i],names[i],contents[i],times[i]]
		except:
			j='空'
			contents.append(j)
			al=[user_ids[i],names[i],contents[i],times[i]]
		all.append(al)
	return all

def save_afile(alls,filename):
    """将数据保存在一个excel"""
    f=xlwt.Workbook()
    sheet1=f.add_sheet(u'sheet1',cell_overwrite_ok=True)
    sheet1.write(0,0,'用户ID')
    sheet1.write(0,1,'用户名')
    sheet1.write(0,2,'评论内容')
    sheet1.write(0,3,'时间')
    i=1
    for all in alls:
        for data in all:
            for j in range(len(data)):
                sheet1.write(i,j,data[j])
            i=i+1
    f.save(r'评论/'+filename+'.xls')#保存路径

def extract(inpath,l):
    """取出一列数据"""
    workbook = openpyxl.load_workbook(inpath)
    sheet = workbook.active  # 获取活动工作表
    numbers=[]
    for row in range(2, sheet.max_row + 1):  # 从第2行开始（跳过表头）
        result = sheet.cell(row=row, column=l+1).value  # openpyxl列索引从1开始
        numbers.append(result)
    return numbers

def save_progress(progress_file, crawled_ids):
    """保存爬取进度"""
    try:
        with open(progress_file, 'w', encoding='utf-8') as f:
            json.dump({'crawled_ids': crawled_ids, 'timestamp': datetime.datetime.now().isoformat()}, f, ensure_ascii=False)
        print(f'进度已保存到 {progress_file}')
    except Exception as e:
        print(f'保存进度失败: {e}')

def load_progress(progress_file):
    """加载爬取进度"""
    try:
        if os.path.exists(progress_file):
            with open(progress_file, 'r', encoding='utf-8') as f:
                progress = json.load(f)
                print(f'从 {progress_file} 加载进度，已爬取 {len(progress.get("crawled_ids", []))} 个微博')
                return set(progress.get('crawled_ids', []))
        else:
            print('未找到进度文件，从头开始爬取')
            return set()
    except Exception as e:
        print(f'加载进度失败: {e}，从头开始爬取')
        return set()

def check_file_exists(bid):
    """检查微博评论文件是否已存在"""
    file_path = os.path.join('评论', f'{bid}.xls')
    return os.path.exists(file_path)

def run(ids, crawled_ids=None, progress_file=None):
	b=ids[0]#bid
	u=str(ids[1]).replace('.0','')#uid
	
	# 检查是否已经爬取过
	if crawled_ids is not None and b in crawled_ids:
		print(f'微博 {b} 已爬取过，跳过')
		return
	
	# 检查文件是否已存在
	if check_file_exists(b):
		print(f'微博 {b} 的文件已存在，跳过')
		if crawled_ids is not None:
			crawled_ids.add(b)
			if progress_file:
				save_progress(progress_file, list(crawled_ids))
		return
	
	alls=[]#每次循环就清空一次
	pa=[]#空列表判定
	url='https://weibo.cn/comment/'+str(b)+'?uid='+str(u)#一个微博的评论首页
	html,page=html_1(url)
	#print(url)
	if page==0:#如果为0，即只有一页数据
		#print('进入页数为0')
		try:
			data_1=body(html)
		except:
			data_1=pa
		alls.append(data_1)#将首页爬取出来
		#print('共计1页,共有'+str(count(alls))+'个数据')
	else:#两页及以上
		#print('进入两页及以上')
		#print('页数为'+str(page))
		for j in range(1,page+1):#从1到page
			if j>=51:
				break
			else:
				url_1=url+'&rl=1'+'&page='+str(j)
				#print(url_1)
				htmls,pages=html_1(url_1)
				alls.append(body(htmls))
			t.sleep(1)
	print('共计'+str(page)+'页,共有'+str(count(alls))+'个数据')
	save_afile(alls,b)

	# 更新爬取进度
	if crawled_ids is not None:
		crawled_ids.add(b)
		if progress_file:
			save_progress(progress_file, list(crawled_ids))

	print('微博号为'+str(b)+'的评论数据文件、保存完毕')

if __name__ == '__main__':
	import os
	
	# 解析命令行参数
	args = parse_arguments()
	
	# 初始化爬虫配置
	if args.config:
		init_crawler(args.config)
	elif args.cookie:
		# 直接传入cookie
		cookie = args.cookie
		headers['Cookie'] = cookie
		print(f"使用命令行传入的cookie，长度: {len(cookie)}")
	else:
		init_crawler()
	
	# 获取当前脚本所在目录的绝对路径
	script_dir = os.path.dirname(os.path.abspath(__file__))
	# 构建正确的文件路径
	fileName = os.path.join(script_dir, '..', '..', 'data', 'concat.xlsx')
	
	# 进度文件路径
	progress_file = os.path.join(script_dir, 'crawler_progress.json')
	
	# 加载爬取进度
	crawled_ids = load_progress(progress_file)
	
	#由于微博限制，只能爬取前五十页的
	#里面的文件是爬取到的正文文件
	bid=extract(fileName,1)#1是bid，2是u_id
	uid=extract(fileName,2)

	ids=[]#将bid和uid匹配并以嵌套列表形式加入ids
	for i,j in zip(bid,uid):
		ids.append([i,j])
	
	print(f'总共需要爬取 {len(ids)} 个微博，已爬取 {len(crawled_ids)} 个，剩余 {len(ids) - len(crawled_ids)} 个')
	
	# 多线程爬取，传入进度信息
	pool = ThreadPool()
	
	# 创建包装函数以传递额外参数
	def run_with_progress(id_pair):
		return run(id_pair, crawled_ids, progress_file)
	
	pool.map(run_with_progress, ids)
	
	print('所有微博评论爬取完成！')
		
		


